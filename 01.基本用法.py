# encoding: utf-8
'''
@author: niko
@contact: simaqingsheng@gmail.com
@file: 01.基本用法.py
@time: 2020/2/4 10:40
@desc:
'''
# 导入urllib中的request模块，用来发送http/https请求
from urllib import request
from bs4 import BeautifulSoup
import pymysql
from openpyxl import Workbook


# 获取数据
# noinspection PyArgumentList
def get_data(i):
    url = "https://search.51job.com/list/030200,000000,0000,00,9,99,java%25E5%25BC%2580%25E5%258F%2591,2," + str(
        i) + ".html"
    # 创建Request对象，指定URL和请求头

    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36'
    }
    req = request.Request(url=url, headers=headers)
    response = request.urlopen(req)

    # print(type(response))  # HTTPResponse类型
    # print(response.getcode())#响应状态码
    # print(response.info())

    if (response.getcode() == 200):
        data = response.read()  # 读取响应结果
        # print(type(data))  # bytes类型
        data = str(data, encoding='gbk')  # 转换为str
        # print(data)
        # 将数据写入文件中
        with open('index.html', mode='w', encoding='gbk') as f:
            f.write(data)


# 解析数据
def parse_data(i):
    get_data(i)
    with open("index.html", mode='r', encoding='gbk') as f:
        html = f.read()

    # 创建BeautifulSoup实例
    bs = BeautifulSoup(html, 'html.parser')

    # 1.find()方法，获取第一个匹配的标签
    # div=bs.find('div')
    # print(div)
    # print(type(div))

    # 2.find_all()方法，获取所有匹配的标签
    # metas=bs.find_all('meta')
    # print(metas)
    # print(metas[0])
    # print(type(metas)) #返回的是list
    # print(bs.find_all(id='hello')) #根据id获取，返回的依然是集合
    # print(bs.find_all(class_='aaa'))#根据class获取，返回的依然是集合

    # 3.select()方法，使用CSS选择器来获取元素
    # print(bs.select('#hello'))
    # print(bs.select('.aaa'))
    # print(bs.select('#world span'))
    # print(bs.select('[title]'))

    # 4.get_text()方法，获取Tag中的文本
    # value = bs.select('#hello')[0].get_text(strip=true)
    # print(len(value))
    # print(value)

    # 获取职位信息
    divs = bs.select("#resultList .el")
    result = []
    # print(divs[0]) #div[0]是表头
    for div in divs[1:]:
        title = div.select('.t1')[0].get_text(strip=True)  # 职位名
        company = div.select('.t2')[0].get_text(strip=True)  # 公司名
        addr = div.select('.t3')[0].get_text(strip=True)  # 工作地点
        salary = div.select('.t4')[0].get_text(strip=True)  # 薪资
        pubDate = div.select('.t5')[0].get_text(strip=True)  # 发布时间
        # print(title, company, addr, salary, pubDate)
        # 将一条记录作为一个字典，存入list中
        row = {
            'title': title,
            'company': company,
            'addr': addr,
            'salary': salary,
            'pubDate': pubDate
        }
        result.append(row)
    return result


# 存储数据到mysql
def save_to_mysql(data):
    config = {
        'host': 'localhost',
        'port': 3306,
        'user': 'root',
        'password': '123456',
        'database': 'python_spider',
        'charset': 'utf8'

    }
    conn = pymysql.connect(**config)
    cursor = conn.cursor()
    sql = '''
    insert into t_job
    (title, company, addr, salary, pubDate) 
    values (%(title)s,%(company)s,%(addr)s,%(salary)s,%(pubDate)s)
    '''
    cursor.executemany(sql, data)
    conn.commit()

    cursor.close()
    conn.close()


# 存储数据到excel
def save_to_excel(data):
    # # 创建工作簿Workbook
    # book=Workbook()
    # # 创建工作表
    # sheet=book.create_sheet('广州Java招聘信息',0)
    #
    # # 向工作表中添加数据
    # sheet.append(['职位名','公司名','工作地点','薪资','发布时间'])
    for item in data:
        row = [item['title'], item['company'], item['addr'], item['salary'], item['pubDate']]
        sheet.append(row)

    # 输出保存
    book.save('广州java招聘信息.xlsx')


def before_save_to_excel():
    # 定义全局变量
    global book, sheet
    book = Workbook()
    sheet = book.create_sheet('广州Java招聘信息', 0)
    sheet.append(['职位名', '公司名', '工作地点', '薪资', '发布时间'])


if __name__ == '__main__':
    # 爬取1500条职位信息数据

    before_save_to_excel();
    for i in range(1, 31):
        # save_to_mysql(parse_data(i))
        save_to_excel(parse_data(i))
